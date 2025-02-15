from django.shortcuts import render,HttpResponse


def homepage(request):
    return render(request, 'Homepage.html')
# Create your views here.
def services(request):
    return render(request, 'Services.html')

# Create your views here.
def contact(request):
    return render(request, 'Contact.html')
# Create your views here.

def destinations(request):
    return render(request, 'Destinations.html')
# Create your views here.
def trip_planing(request):
    return render(request, 'trip_planing.html')

def custom_itineraries(request):
    return render(request, 'custom_itineraries.html')

def group_discounts(request):
    return render(request, 'group_discounts.html')

def travel_insurance(request):
    return render(request, 'travel_insurance.html')

def visa_assistance(request):
    return render(request, 'visa_assistance.html')

def customer_support(request):
    return render(request,"customer.html")

def paris(request):
    return render(request,"Paris.html")

def rome(request):
    return render(request,"rome.html")

def bali(request):
    return render(request,"bali.html")

def capetown(request):
    return render(request,"capetown.html")

def newyork(request):
    return render(request,"newyork.html")

def sydney(request):
    return render(request,"sydney.html")

def tokyo(request):
    return render(request,"tokyo.html")

def dubai(request):
    return render(request,"dubai.html")


def book(request):
    destination = request.GET.get('destination', 'Unknown')
    return render(request, 'booking.html', {'destination': destination})

def effiel(requrst):
    return render(requrst, 'effiel.html')

def louvre(requrst):
    return render(requrst, 'louvre.html')


def login(request):
    return render(request,"login.html")



from django.http import JsonResponse
from .utils import save_user_to_excel
from django.shortcuts import render



def signup(request):
    if request.method == "POST":
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        password = request.POST.get("password")

        if save_user_to_excel(email, phone, password):
            return JsonResponse({"message": "Signup successful!"}, status=201)
        else:
            return JsonResponse({"message": "Error saving user data."}, status=500)
    
    # If GET request, render the signup page
    return render(request, 'signup.html')

from django.http import JsonResponse
from .utils import save_contact_to_excel
from django.shortcuts import render

def contact(request):
    if request.method == "POST":
        name=request.POST.get("name")
        email=request.POST.get("email")
        message=request.POST.get("message")
        if save_contact_to_excel(name,email,message):
            return JsonResponse({"message": "Signup successful!"}, status=201)
        else:
            return JsonResponse({"message": "Error saving user data."}, status=500)
    
    return render(request,"Contact.html")
# views.py
from django.http import JsonResponse
from .utils import save_contact_to_excel
from django.shortcuts import render

def contact(request):
    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        message = request.POST.get("message")
        success = save_contact_to_excel(name, email, message)
        if success:
            return JsonResponse({"message": "Message sent successfully!"}, status=201)
        else:
            return JsonResponse({"message": "Error saving user data."}, status=500)
    
    return render(request, "Contact.html")


def explore_destinations(request):
    return render(request, 'explore_destinations.html')

import os
import openpyxl
from django.conf import settings
from django.core.mail import EmailMessage
from django.shortcuts import render
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.template.loader import get_template
from io import BytesIO

def submit_booking(request):
    if request.method == 'POST':
        # Extract booking details
        name = request.POST.get('name')
        email = request.POST.get('email')
        destination = request.POST.get('destination')
        date = request.POST.get('date')
        guests = request.POST.get('guests')
        notes = request.POST.get('notes')

        # Save booking to Excel
        booking_saved = save_booking_to_excel(name, email, destination, date, guests, notes)
        if not booking_saved:
            return HttpResponse("Failed to save booking. Please try again.")

        # Generate ticket PDF
        ticket_path = generate_ticket_pdf(name, destination, date, guests, notes)
        if not ticket_path:
            return HttpResponse("Error generating the ticket. Please try again later.")

        # Send ticket via email
        email_sent = send_ticket_email(email, name, ticket_path)
        if not email_sent:
            return HttpResponse("Booking saved, but failed to send email.")

        return HttpResponse(f"Thank you, {name}! Your trip to {destination} has been booked. Check your email for the ticket.")
    
    return HttpResponse("Invalid request.")

from openpyxl import Workbook
import tempfile
from datetime import datetime

import os
from openpyxl import Workbook
from datetime import datetime

import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def save_booking_to_excel(name, email, destination, date, guests, notes):
    try:
        # Define the path where the Excel file will be saved
        excel_path = 'travel_app/storage/bookings.xlsx'

        # Check if the file exists
        if os.path.exists(excel_path):
            # Load the existing workbook
            wb = load_workbook(excel_path)
            ws = wb.active
        else:
            # Create a new workbook if the file doesn't exist
            wb = Workbook()
            ws = wb.active
            # Add headers if creating a new file
            ws.append(['Name', 'Email', 'Destination', 'Travel Date', 'Guests', 'Notes', 'Booking Date'])

        # Append booking details with the current timestamp
        ws.append([name, email, destination, date, guests, notes, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

        # Save the workbook with the updated data
        wb.save(excel_path)

        print(f"Booking saved to Excel file: {excel_path}")
        return True
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False


def generate_ticket_pdf(name, destination, date, guests, notes):
    try:
        template = get_template('ticket_template.html')
        context = {'name': name, 'destination': destination, 'date': date, 'guests': guests, 'notes': notes}
        html = template.render(context)
        ticket_path = os.path.join(settings.MEDIA_ROOT, f"{name}_ticket.pdf")
        with open(ticket_path, "wb") as pdf_file:
            pisa_status = pisa.CreatePDF(html, dest=pdf_file)
        if pisa_status.err:
            raise Exception("PDF generation failed")
        return ticket_path
    except Exception as e:
        print(f"Error generating PDF: {e}")
        return None

import os
from django.conf import settings

def send_ticket_email(email, name, ticket_path):
    try:
        subject = "Your Trip Booking Confirmation"
        body = f"Hi {name},\n\nThank you for booking your trip with us! Please find your ticket attached."

        email_message = EmailMessage(
            subject,
            body,
            settings.EMAIL_HOST_USER,
            [email],
        )

        # Attach the ticket file if it exists
        if os.path.exists(ticket_path):
            email_message.attach_file(ticket_path)

        # Send the email with fail_silently=False for debugging
        email_message.send(fail_silently=False)

        print("Email sent successfully.")
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False



def test_file_write(request):
    try:
        test_file_path = os.path.join(settings.MEDIA_ROOT, 'test.txt')
        with open(test_file_path, 'w') as f:
            f.write('test')
        return HttpResponse("File write successful")
    except Exception as e:
        return HttpResponse(f"Error writing file: {e}")
    

