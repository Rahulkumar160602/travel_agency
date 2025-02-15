import openpyxl
import os
from datetime import datetime
from django.conf import settings

def save_user_to_excel(email, phone, password):
    # Absolute path to the user_data.xlsx file
    file_path = os.path.join(settings.BASE_DIR, "travel_app", "storage", "user_data.xlsx")
    try:
        # Load the Excel file
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Add the user's data to the sheet
        signup_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([email, phone, password, signup_date])

        # Save the changes
        workbook.save(file_path)
        return True
    except Exception as e:
        print(f"Error saving user to Excel: {e}")
        return False

import os
import openpyxl

# Get the base directory of the app
from datetime import datetime

def save_booking_to_excel(name, email, destination, date, guests, notes):
    # Absolute path to the bookings.xlsx file
    file_path = os.path.join(settings.BASE_DIR, "travel_app", "storage", "bookings.xlsx")

    try:
        # Check if the Excel file exists
        if not os.path.exists(file_path):
            # Create a new Excel file with headers
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Email", "Destination", "Travel Date", "Guests", "Notes", "Booking Time"])
        else:
            # Load the existing Excel file
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

        # Add the booking data with a timestamp
        booking_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([name, email, destination, date, guests, notes, booking_time])

        # Save the changes
        workbook.save(file_path)
        print("Booking data saved successfully!")  # Debug message
        return True
    except Exception as e:
        print(f"Error saving booking to Excel: {e}")
        return False

import openpyxl
import os
from datetime import datetime
from django.conf import settings

def save_contact_to_excel(name, email, message):
    # Absolute path to the user_data.xlsx file
    file_path = os.path.join(settings.BASE_DIR, "travel_app", "storage", "contact.xlsx")
    
    # Log the file path being used
    print(f"File Path: {file_path}")
    
    try:
        # If the file doesn't exist, create a new one and add headers
        if not os.path.exists(file_path):
            print("File does not exist. Creating a new file.")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Email", "Message", "Signup Date"])  # Add headers
            workbook.save(file_path)

        # If the file is corrupted, delete it and create a new one
        if file_path.endswith(".xlsx"):
            with open(file_path, 'rb') as f:
                if not f.read(4) == b"PK\x03\x04":
                    print("File is corrupted. Re-creating the file.")
                    os.remove(file_path)  # Remove the corrupted file
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet.append(["Name", "Email", "Message", "Signup Date"])  # Add headers
                    workbook.save(file_path)

        # Load the Excel file
        print("Loading the workbook.")
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Add the user's data to the sheet
        signup_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"Appending data: {name}, {email}, {message}, {signup_date}")
        sheet.append([name, email, message, signup_date])

        # Save the changes
        print("Saving the workbook.")
        workbook.save(file_path)
        return True
    except Exception as e:
        print(f"Error saving user to Excel: {e}")  # Print the full exception
        return False
